from fastapi import APIRouter, Depends, HTTPException, status, Query, Request, BackgroundTasks
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from datetime import date, datetime
from pydantic import BaseModel, EmailStr, Field, AliasChoices
from app.db.session import get_db
from app.db.models.client import Client, Director, BusinessType, ClientStatus, ServiceType
from app.db.models.service import Service
from app.db.models.user import User, UserRole
from app.db.models.organization import Organization
from app.core.security import get_password_hash, validate_password, validate_client_password, generate_secure_password, encrypt_password, decrypt_password
from app.core.email_service import send_login_credentials_email
from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import logging

router = APIRouter()

# Include email config router
from app.api.v1.client import email_config
router.include_router(email_config.router)


# Pydantic schemas
class DirectorCreate(BaseModel):
    director_name: str
    email: Optional[EmailStr] = None
    phone_number: Optional[str] = None
    # Field name is "designation"; we still accept old key "resignation" for backward compatibility
    designation: Optional[str] = Field(
        default=None,
        validation_alias=AliasChoices("designation", "resignation"),
    )
    din: Optional[str] = None
    pan: Optional[str] = None
    aadhaar: Optional[str] = None


class DirectorResponse(BaseModel):
    id: int
    director_name: str
    email: Optional[str] = None
    phone_number: Optional[str] = None
    designation: Optional[str] = None
    din: Optional[str] = None
    pan: Optional[str] = None
    aadhaar: Optional[str] = None
    
    class Config:
        from_attributes = True


class ServiceResponse(BaseModel):
    id: int
    name: str
    is_custom: bool
    
    class Config:
        from_attributes = True


class ClientCreate(BaseModel):
    client_name: str
    email: Optional[EmailStr] = None
    phone_number: str
    company_name: str
    business_type: BusinessType
    pan_number: Optional[str] = None
    gst_number: Optional[str] = None
    status: ClientStatus = ClientStatus.ACTIVE
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = None
    pin_code: Optional[str] = None
    onboard_date: Optional[date] = None
    follow_date: Optional[date] = None
    additional_notes: Optional[str] = None
    service_ids: List[int] = []
    directors: List[DirectorCreate] = []
    # Login credentials (optional - admin provides these)
    login_email: Optional[EmailStr] = None
    login_password: Optional[str] = None


class ClientUpdate(BaseModel):
    client_name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone_number: Optional[str] = None
    company_name: Optional[str] = None
    business_type: Optional[BusinessType] = None
    pan_number: Optional[str] = None
    gst_number: Optional[str] = None
    status: Optional[ClientStatus] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = None
    pin_code: Optional[str] = None
    onboard_date: Optional[date] = None
    follow_date: Optional[date] = None
    additional_notes: Optional[str] = None
    service_ids: Optional[List[int]] = None
    directors: Optional[List[DirectorCreate]] = None
    # Login credentials (optional - admin can update these)
    login_email: Optional[EmailStr] = None
    login_password: Optional[str] = None


class ClientResponse(BaseModel):
    id: int
    client_name: str
    email: Optional[str] = None
    phone_number: str
    company_name: str
    business_type: BusinessType
    pan_number: Optional[str] = None
    gst_number: Optional[str] = None
    status: ClientStatus
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = None
    pin_code: Optional[str] = None
    onboard_date: date
    follow_date: Optional[date] = None
    additional_notes: Optional[str] = None
    services: List[ServiceResponse] = []
    directors: List[DirectorResponse] = []
    created_at: datetime
    updated_at: Optional[datetime] = None
    # Login credentials (if client has login account)
    login_email: Optional[str] = None  # Email used for login (from User.email)
    user_id: Optional[int] = None  # ID of the User account (indicates login account exists)
    login_password: Optional[str] = None  # Password (only returned on create/update, not on GET)
    
    class Config:
        from_attributes = True


class AdminResponse(BaseModel):
    id: int
    email: str
    full_name: Optional[str] = None
    phone: Optional[str] = None
    role: str
    
    class Config:
        from_attributes = True


class OrganizationDetailResponse(BaseModel):
    id: int
    name: str
    city: Optional[str] = None
    state: Optional[str] = None
    country: Optional[str] = None
    pincode: Optional[str] = None
    
    class Config:
        from_attributes = True


class ClientsListResponse(BaseModel):
    admin: AdminResponse
    organization: OrganizationDetailResponse
    clients: List[ClientResponse]
    total: int
    skip: int
    limit: int


class ServiceCreate(BaseModel):
    name: str


def initialize_default_services(db: Session):
    """Initialize default services if they don't exist"""
    for service_type in ServiceType:
        existing_service = db.query(Service).filter(Service.name == service_type.value).first()
        if not existing_service:
            db_service = Service(name=service_type.value, is_custom=False)
            db.add(db_service)
    db.commit()


def build_client_response(client: Client, login_password: Optional[str] = None) -> ClientResponse:
    """
    Helper function to build ClientResponse from Client model.
    Includes login credentials (login_email, user_id) if client has a user account.
    
    Args:
        client: Client model instance
        login_password: Optional plain password to include in response (only for create/update operations).
                       If None and client has a user account, will try to decrypt stored password.
    """
    # If login_password not provided but client has user account, try to decrypt stored password
    if login_password is None and client.user_id and client.user:
        login_password = decrypt_password(client.user.encrypted_plain_password)
    # Build base response dict
    client_dict = {
        "id": client.id,
        "client_name": client.client_name,
        "email": client.email,
        "phone_number": client.phone_number,
        "company_name": client.company_name,
        "business_type": client.business_type,
        "pan_number": client.pan_number,
        "gst_number": client.gst_number,
        "status": client.status,
        "address": client.address,
        "city": client.city,
        "state": client.state,
        "country": client.country,
        "pin_code": client.pin_code,
        "onboard_date": client.onboard_date,
        "follow_date": client.follow_date,
        "additional_notes": client.additional_notes,
        "services": client.services,
        "directors": client.directors,
        "created_at": client.created_at,
        "updated_at": client.updated_at,
        "login_email": None,
        "user_id": client.user_id,
        "login_password": login_password  # Only included if provided (create/update operations)
    }
    
    # If client has a user account, get the login email
    if client.user_id and client.user:
        client_dict["login_email"] = client.user.email
    
    return ClientResponse(**client_dict)


@router.post("/", response_model=ClientResponse, status_code=status.HTTP_201_CREATED)
def create_client(
    client: ClientCreate,
    request: Request,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    Create a new client with directors and services.
    If login_email and login_password are provided, creates a User account with CLIENT role
    so the client can login to the system.
    """
    # Get org_id from request state (set by TenantMiddleware)
    org_id = request.state.org_id
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required. Please provide a valid access token."
        )
    
    # Initialize default services
    initialize_default_services(db)
    
    # Validate services
    if client.service_ids:
        services = db.query(Service).filter(Service.id.in_(client.service_ids)).all()
        if len(services) != len(client.service_ids):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="One or more service IDs are invalid"
            )
    
    # Handle login credentials if provided
    user_id = None
    plain_password = None  # Store plain password to return in response
    password_was_generated = False  # Track if password was auto-generated
    
    if client.login_email:
        # Check if user with this email already exists
        existing_user = db.query(User).filter(User.email == client.login_email).first()
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="User with this email already exists"
            )
        
        # Determine password to use
        if client.login_password:
            # Use provided password
            plain_password = client.login_password
            
            # Validate password strength (relaxed for client login)
            is_valid, error_message = validate_client_password(client.login_password)
            if not is_valid:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid password: {error_message}"
                )
        else:
            # Generate secure password automatically
            plain_password = generate_secure_password(length=12)
            password_was_generated = True
        
        # Create user account for client
        db_user = User(
            email=client.login_email,
            hashed_password=get_password_hash(plain_password),
            encrypted_plain_password=encrypt_password(plain_password),  # Store encrypted plain password
            full_name=client.client_name,
            phone=client.phone_number,
            org_id=org_id,
            role=UserRole.CLIENT
        )
        db.add(db_user)
        db.flush()  # Flush to get the user_id
        user_id = db_user.id
    
    # Create client
    db_client = Client(
        client_name=client.client_name,
        email=client.email,
        phone_number=client.phone_number,
        company_name=client.company_name,
        business_type=client.business_type,
        pan_number=client.pan_number,
        gst_number=client.gst_number,
        status=client.status,
        address=client.address,
        city=client.city,
        state=client.state,
        country=client.country,
        pin_code=client.pin_code,
        onboard_date=client.onboard_date or date.today(),
        follow_date=client.follow_date,
        additional_notes=client.additional_notes,
        user_id=user_id,
        org_id=org_id
    )
    
    # Add services
    if client.service_ids:
        db_client.services = services
    
    db.add(db_client)
    db.flush()  # Flush to get the client ID
    
    # Add directors
    for director_data in client.directors:
        db_director = Director(
            client_id=db_client.id,
            director_name=director_data.director_name,
            email=director_data.email,
            phone_number=director_data.phone_number,
            designation=director_data.designation,
            din=director_data.din,
            pan=director_data.pan,
            aadhaar=director_data.aadhaar
        )
        db.add(db_director)
    
    db.commit()
    db.refresh(db_client)
    # Reload with user relationship to include login credentials in response
    db_client = db.query(Client).options(joinedload(Client.user)).filter(Client.id == db_client.id).first()
    
    # Send login credentials email to client if login credentials were provided (in background)
    if plain_password and client.login_email:
        # Get organization name for email
        org = db.query(Organization).filter(Organization.id == org_id).first()
        org_name = org.name if org else None
        
        background_tasks.add_task(
            send_login_credentials_email,
            recipient_email=client.login_email,
            recipient_name=client.client_name,
            login_email=client.login_email,
            password=plain_password,
            role="client",
            organization_name=org_name
        )
    
    # Return response with plain password (only on create)
    return build_client_response(db_client, login_password=plain_password)


def parse_status_filter(value: Optional[str]) -> Optional[ClientStatus]:
    """Convert empty string to None and parse ClientStatus enum"""
    if not value or not value.strip():
        return None
    try:
        return ClientStatus(value.strip().lower())
    except ValueError:
        return None


@router.get("/", response_model=ClientsListResponse)
def get_clients(
    request: Request,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    search: Optional[str] = Query(default=None, description="Search term for client name, company name, email, or phone"),
    status_filter: Optional[str] = Query(default=None, description="Filter by status: active, inactive, or terminated"),
    db: Session = Depends(get_db)
):
    """
    Get all clients with admin details, organization details, and company details.
    Returns a comprehensive response including admin info, organization info, and list of clients.
    """
    # Get org_id from request state (set by TenantMiddleware)
    org_id = request.state.org_id
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required. Please provide a valid access token."
        )
    
    # Get organization details
    organization = db.query(Organization).filter(Organization.id == org_id).first()
    if not organization:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Organization not found"
        )
    
    # Get admin user for this organization
    admin = db.query(User).filter(
        User.org_id == org_id,
        User.role == UserRole.ADMIN
    ).first()
    
    if not admin:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Admin user not found for this organization"
        )
    
    # Get clients query - filter by org_id and eager load user relationship
    query = db.query(Client).options(joinedload(Client.user)).filter(Client.org_id == org_id)
    
    # Apply search filter (handle empty string as None)
    if search and search.strip():
        search_term = f"%{search.strip()}%"
        query = query.filter(
            or_(
                Client.client_name.ilike(search_term),
                Client.company_name.ilike(search_term),
                Client.email.ilike(search_term),
                Client.phone_number.ilike(search_term)
            )
        )
    
    # Apply status filter (parse enum from string, handle empty strings)
    parsed_status = parse_status_filter(status_filter)
    if parsed_status:
        query = query.filter(Client.status == parsed_status)
    
    # Get total count before pagination
    total = query.count()
    
    # Get paginated clients with user relationship loaded
    clients = query.offset(skip).limit(limit).all()
    
    # Build ClientResponse objects with login credentials
    client_responses = [build_client_response(client) for client in clients]
    
    return ClientsListResponse(
        admin=AdminResponse(
            id=admin.id,
            email=admin.email,
            full_name=admin.full_name,
            phone=admin.phone,
            role=admin.role.value
        ),
        organization=OrganizationDetailResponse(
            id=organization.id,
            name=organization.name,
            city=organization.city,
            state=organization.state,
            country=organization.country,
            pincode=organization.pincode
        ),
        clients=client_responses,
        total=total,
        skip=skip,
        limit=limit
    )


@router.get("/{client_id}", response_model=ClientResponse)
def get_client(
    client_id: int,
    db: Session = Depends(get_db)
):
    """
    Get a specific client by ID.
    Returns login credentials (login_email, user_id) if the client has a login account.
    Note: login_password is not returned for security reasons.
    """
    client = db.query(Client).options(joinedload(Client.user)).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Client not found"
        )
    
    # Build response with login credentials if available
    return build_client_response(client)


@router.put("/{client_id}", response_model=ClientResponse)
def update_client(
    client_id: int,
    client_update: ClientUpdate,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Update a client.
    If login_email and/or login_password are provided, updates or creates the client's user account.
    """
    # Get org_id from request state (set by TenantMiddleware)
    org_id = request.state.org_id
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required. Please provide a valid access token."
        )
    
    db_client = db.query(Client).filter(Client.id == client_id).first()
    if not db_client:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Client not found"
        )
    
    # Handle login credentials update
    login_email = client_update.login_email
    login_password = client_update.login_password
    plain_password = None  # Store plain password to return in response if updated/generated
    password_was_generated = False  # Track if password was auto-generated
    
    if login_email is not None or login_password is not None:
        # If client already has a user account
        if db_client.user_id:
            db_user = db.query(User).filter(User.id == db_client.user_id).first()
            if db_user:
                # Update email if provided
                if login_email is not None and login_email != db_user.email:
                    # Check if new email already exists
                    existing_user = db.query(User).filter(
                        User.email == login_email,
                        User.id != db_user.id
                    ).first()
                    if existing_user:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail="User with this email already exists"
                        )
                    db_user.email = login_email
                
                # Update password if provided, or generate new one if only email is provided
                if login_password:
                    # Use provided password
                    plain_password = login_password
                    is_valid, error_message = validate_client_password(login_password)
                    if not is_valid:
                        raise HTTPException(
                            status_code=status.HTTP_400_BAD_REQUEST,
                            detail=f"Invalid password: {error_message}"
                        )
                    db_user.hashed_password = get_password_hash(login_password)
                    db_user.encrypted_plain_password = encrypt_password(login_password)  # Store encrypted plain password
                elif login_email is not None:
                    # If only login_email provided (no password), generate a new password
                    plain_password = generate_secure_password(length=12)
                    password_was_generated = True
                    db_user.hashed_password = get_password_hash(plain_password)
                    db_user.encrypted_plain_password = encrypt_password(plain_password)  # Store encrypted plain password
                
                # Update user details from client
                if client_update.client_name:
                    db_user.full_name = client_update.client_name
                if client_update.phone_number:
                    db_user.phone = client_update.phone_number
        else:
            # Client doesn't have a user account yet - create one
            if not login_email:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="login_email is required to create a new login account"
                )
            
            # Determine password to use
            if login_password:
                # Use provided password
                plain_password = login_password
                
                # Validate password strength (relaxed for client login)
                is_valid, error_message = validate_client_password(login_password)
                if not is_valid:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Invalid password: {error_message}"
                    )
            else:
                # Generate secure password automatically
                plain_password = generate_secure_password(length=12)
                password_was_generated = True
            
            # Check if user with this email already exists
            existing_user = db.query(User).filter(User.email == login_email).first()
            if existing_user:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="User with this email already exists"
                )
            
            # Create user account for client
            db_user = User(
                email=login_email,
                hashed_password=get_password_hash(plain_password),
                encrypted_plain_password=encrypt_password(plain_password),  # Store encrypted plain password
                full_name=client_update.client_name or db_client.client_name,
                phone=client_update.phone_number or db_client.phone_number,
                org_id=org_id,
                role=UserRole.CLIENT
            )
            db.add(db_user)
            db.flush()  # Flush to get the user_id
            db_client.user_id = db_user.id
    
    # Update client fields (exclude login credentials from client fields)
    update_data = client_update.dict(
        exclude_unset=True, 
        exclude={'service_ids', 'directors', 'login_email', 'login_password'}
    )
    for field, value in update_data.items():
        setattr(db_client, field, value)
    
    # Update services if provided
    if client_update.service_ids is not None:
        services = db.query(Service).filter(Service.id.in_(client_update.service_ids)).all()
        if len(services) != len(client_update.service_ids):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="One or more service IDs are invalid"
            )
        db_client.services = services
    
    # Update directors if provided
    if client_update.directors is not None:
        # Delete existing directors
        db.query(Director).filter(Director.client_id == client_id).delete()
        # Add new directors
        for director_data in client_update.directors:
            db_director = Director(
                client_id=db_client.id,
                director_name=director_data.director_name,
                email=director_data.email,
                phone_number=director_data.phone_number,
                designation=director_data.designation,
                din=director_data.din,
                pan=director_data.pan,
                aadhaar=director_data.aadhaar
            )
            db.add(db_director)
    
    db.commit()
    db.refresh(db_client)
    # Reload with user relationship to include login credentials in response
    db_client = db.query(Client).options(joinedload(Client.user)).filter(Client.id == db_client.id).first()
    # Return response with plain password (only on create)
    return build_client_response(db_client, login_password=plain_password)


@router.delete("/{client_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_client(
    client_id: int,
    db: Session = Depends(get_db)
):
    """Delete a client"""
    db_client = db.query(Client).filter(Client.id == client_id).first()
    if not db_client:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Client not found"
        )
    
    db.delete(db_client)
    db.commit()
    return None


@router.get("/export/excel")
def export_clients_to_excel(
    db: Session = Depends(get_db)
):
    """Export all clients to Excel"""
    clients = db.query(Client).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "ID", "Client Name", "Email", "Phone Number", "Company Name", "Business Type",
        "PAN Number", "GST Number", "Status", "Address", "City", "State", "Country",
        "Pin Code", "Onboard Date", "Follow Date", "Additional Notes", "Services", "Directors"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, client in enumerate(clients, 2):
        services_str = ", ".join([s.name for s in client.services])
        directors_str = ", ".join([d.director_name for d in client.directors])
        
        ws.cell(row=row_num, column=1, value=client.id)
        ws.cell(row=row_num, column=2, value=client.client_name)
        ws.cell(row=row_num, column=3, value=client.email or "")
        ws.cell(row=row_num, column=4, value=client.phone_number)
        ws.cell(row=row_num, column=5, value=client.company_name)
        ws.cell(row=row_num, column=6, value=client.business_type.value)
        ws.cell(row=row_num, column=7, value=client.pan_number or "")
        ws.cell(row=row_num, column=8, value=client.gst_number or "")
        ws.cell(row=row_num, column=9, value=client.status.value)
        ws.cell(row=row_num, column=10, value=client.address or "")
        ws.cell(row=row_num, column=11, value=client.city or "")
        ws.cell(row=row_num, column=12, value=client.state or "")
        ws.cell(row=row_num, column=13, value=client.country or "")
        ws.cell(row=row_num, column=14, value=client.pin_code or "")
        ws.cell(row=row_num, column=15, value=client.onboard_date.strftime("%Y-%m-%d") if client.onboard_date else "")
        ws.cell(row=row_num, column=16, value=client.follow_date.strftime("%Y-%m-%d") if client.follow_date else "")
        ws.cell(row=row_num, column=17, value=client.additional_notes or "")
        ws.cell(row=row_num, column=18, value=services_str)
        ws.cell(row=row_num, column=19, value=directors_str)
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clients_export.xlsx"}
    )


# Service management endpoints
@router.get("/services/", response_model=List[ServiceResponse])
def get_services(
    db: Session = Depends(get_db)
):
    """Get all available services"""
    initialize_default_services(db)
    services = db.query(Service).order_by(Service.is_custom, Service.name).all()
    return services


@router.post("/services/", response_model=ServiceResponse, status_code=status.HTTP_201_CREATED)
def create_custom_service(
    service: ServiceCreate,
    db: Session = Depends(get_db)
):
    """Create a custom service"""
    # Check if service already exists
    existing_service = db.query(Service).filter(Service.name == service.name).first()
    if existing_service:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Service with this name already exists"
        )
    
    db_service = Service(name=service.name, is_custom=True)
    db.add(db_service)
    db.commit()
    db.refresh(db_service)
    return db_service


# Enum endpoints for frontend
@router.get("/enums/status/")
def get_status_enum():
    """Get all available client status values"""
    return {
        "values": [status.value for status in ClientStatus],
        "enum": {status.name: status.value for status in ClientStatus}
    }


@router.get("/enums/business-type/")
def get_business_type_enum():
    """Get all available business type values"""
    return {
        "values": [bt.value for bt in BusinessType],
        "enum": {bt.name: bt.value for bt in BusinessType}
    }


@router.get("/enums/service-type/")
def get_service_type_enum():
    """Get all available default service type values"""
    return {
        "values": [st.value for st in ServiceType],
        "enum": {st.name: st.value for st in ServiceType}
    }


# Director management endpoints
@router.post("/{client_id}/directors/", response_model=DirectorResponse, status_code=status.HTTP_201_CREATED)
def add_director(
    client_id: int,
    director: DirectorCreate,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Add a new director to an existing client.
    """
    # Get org_id from request state (set by TenantMiddleware)
    org_id = request.state.org_id
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required. Please provide a valid access token."
        )
    
    # Verify client exists and belongs to the organization
    client = db.query(Client).filter(
        Client.id == client_id,
        Client.org_id == org_id
    ).first()
    
    if not client:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Client not found"
        )
    
    # Create director
    db_director = Director(
        client_id=client_id,
        director_name=director.director_name,
        email=director.email,
        phone_number=director.phone_number,
        designation=director.designation,
        din=director.din,
        pan=director.pan,
        aadhaar=director.aadhaar
    )
    db.add(db_director)
    db.commit()
    db.refresh(db_director)
    
    return db_director


@router.get("/{client_id}/directors/", response_model=List[DirectorResponse])
def get_directors(
    client_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Get all directors for a specific client.
    """
    # Get org_id from request state (set by TenantMiddleware)
    org_id = request.state.org_id
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required. Please provide a valid access token."
        )
    
    # Verify client exists and belongs to the organization
    client = db.query(Client).filter(
        Client.id == client_id,
        Client.org_id == org_id
    ).first()
    
    if not client:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Client not found"
        )
    
    # Get all directors for this client
    directors = db.query(Director).filter(Director.client_id == client_id).all()
    return directors


@router.put("/directors/{director_id}", response_model=DirectorResponse)
def update_director(
    director_id: int,
    director: DirectorCreate,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Update an existing director.
    """
    # Get org_id from request state (set by TenantMiddleware)
    org_id = request.state.org_id
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required. Please provide a valid access token."
        )
    
    # Get director and verify it belongs to a client in this organization
    db_director = db.query(Director).join(Client).filter(
        Director.id == director_id,
        Client.org_id == org_id
    ).first()
    
    if not db_director:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Director not found"
        )
    
    # Update director fields
    db_director.director_name = director.director_name
    db_director.email = director.email
    db_director.phone_number = director.phone_number
    db_director.designation = director.designation
    db_director.din = director.din
    db_director.pan = director.pan
    db_director.aadhaar = director.aadhaar
    
    db.commit()
    db.refresh(db_director)
    
    return db_director


@router.delete("/directors/{director_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_director(
    director_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Delete a director.
    """
    # Get org_id from request state (set by TenantMiddleware)
    org_id = request.state.org_id
    if not org_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Authentication required. Please provide a valid access token."
        )
    
    # Get director and verify it belongs to a client in this organization
    db_director = db.query(Director).join(Client).filter(
        Director.id == director_id,
        Client.org_id == org_id
    ).first()
    
    if not db_director:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Director not found"
        )
    
    db.delete(db_director)
    db.commit()
    return None

